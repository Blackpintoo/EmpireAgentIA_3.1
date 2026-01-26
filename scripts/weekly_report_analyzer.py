#!/usr/bin/env python
"""
Weekly Trading Report Analyzer for EmpireAgentIA_3
Analyzes trading performance over the last 7 days and generates a detailed report.

Usage:
    python scripts/weekly_report_analyzer.py
    python scripts/weekly_report_analyzer.py --days 14
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

# Add project root to path
ROOT = Path(__file__).parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import load_workbook

LOCAL_TZ = ZoneInfo("Europe/Zurich")
UTC = timezone.utc

# ============================================================================
# DATA LOADING FUNCTIONS
# ============================================================================

def load_mt5_history_from_excel(filepath: Path, start_date: datetime, end_date: datetime) -> pd.DataFrame:
    """Load MT5 trade history from Excel report."""
    if not filepath.exists():
        print(f"[WARNING] Excel file not found: {filepath}")
        return pd.DataFrame()

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # Find header row (contains "Heure" and "Position")
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == 'Heure' and len(row) > 1 and row[1] == 'Position':
            header_row = idx
            break

    if header_row is None:
        print("[WARNING] Could not find header row in Excel")
        return pd.DataFrame()

    # Extract data
    data = []
    headers = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if idx == header_row:
            headers = [str(h) if h else f'col_{i}' for i, h in enumerate(row)]
        elif idx > header_row and row and row[0]:
            data.append(list(row))

    if not headers or not data:
        print("[WARNING] No data found in Excel")
        return pd.DataFrame()

    df = pd.DataFrame(data, columns=headers)
    print(f"[DEBUG] Excel columns: {list(df.columns)}")

    # Parse dates
    def parse_date(x):
        if pd.isna(x) or x is None:
            return None
        try:
            return datetime.strptime(str(x), "%Y.%m.%d %H:%M:%S").replace(tzinfo=LOCAL_TZ)
        except:
            return None

    df['open_time'] = df['Heure'].apply(parse_date)
    df['close_time'] = df['Heure.1'].apply(parse_date)

    # Convert numeric columns
    for col in ['Volume', 'Prix', 'Prix.1', 'Commission', 'Echange', 'Profit']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Filter by date range
    df = df[df['close_time'].notna()]
    df = df[(df['close_time'] >= start_date) & (df['close_time'] <= end_date)]

    # Rename columns for consistency
    df = df.rename(columns={
        'Symbole': 'symbol',
        'Type': 'type',
        'Volume': 'volume',
        'Prix': 'open_price',
        'Prix.1': 'close_price',
        'Commission': 'commission',
        'Echange': 'swap',
        'Profit': 'profit',
        'Position': 'position_id'
    })

    return df


def load_journal_trades(data_dir: Path, start_date: datetime, end_date: datetime) -> List[Dict]:
    """Load trades from daily JSONL journal files."""
    trades = []
    journal_dir = data_dir / "journal"

    if not journal_dir.exists():
        return trades

    current = start_date.date()
    end = end_date.date()

    while current <= end:
        file_path = journal_dir / f"trades_{current.isoformat()}.jsonl"
        if file_path.exists():
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line:
                        try:
                            trade = json.loads(line)
                            trades.append(trade)
                        except json.JSONDecodeError:
                            pass
        current += timedelta(days=1)

    return trades


def load_audit_logs(logs_dir: Path, start_date: datetime, end_date: datetime) -> List[Dict]:
    """Load audit logs from JSONL file, filtering by date range."""
    audit_file = logs_dir / "audit_trades.jsonl"
    logs = []

    if not audit_file.exists():
        return logs

    # Read only last portion of file to avoid memory issues
    with open(audit_file, 'r', encoding='utf-8') as f:
        # Read last 10MB
        f.seek(0, 2)
        file_size = f.tell()
        read_size = min(10 * 1024 * 1024, file_size)
        f.seek(max(0, file_size - read_size))

        # Skip partial first line
        if f.tell() > 0:
            f.readline()

        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                entry = json.loads(line)
                ts = entry.get('ts', '')
                if ts:
                    entry_dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
                    if start_date <= entry_dt <= end_date:
                        logs.append(entry)
            except (json.JSONDecodeError, ValueError):
                pass

    return logs


def load_trade_outcomes(data_dir: Path, start_date: datetime, end_date: datetime) -> pd.DataFrame:
    """Load trade outcomes with exit_type from CSV."""
    outcomes_file = data_dir / "trade_outcomes.csv"

    if not outcomes_file.exists():
        return pd.DataFrame()

    try:
        df = pd.read_csv(outcomes_file)

        # Parse datetime
        if 'close_time' in df.columns:
            df['close_time'] = pd.to_datetime(df['close_time'], errors='coerce')
            df = df[df['close_time'].notna()]
            df = df[(df['close_time'] >= start_date) & (df['close_time'] <= end_date)]

        # Ensure exit_type column exists
        if 'exit_type' not in df.columns:
            df['exit_type'] = 'unknown'

        return df
    except Exception as e:
        print(f"[WARNING] Could not load trade_outcomes.csv: {e}")
        return pd.DataFrame()


def analyze_by_exit_type(df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """Analyze performance by exit type (tp, sl, manual, etc.)."""
    if df.empty or 'exit_type' not in df.columns:
        return {}

    results = {}
    for exit_type, group in df.groupby('exit_type'):
        total = len(group)
        winners = len(group[group['profit'] > 0])
        win_rate = (winners / total * 100) if total > 0 else 0
        total_pnl = group['profit'].sum()
        avg_pnl = group['profit'].mean()

        results[exit_type] = {
            'trades': total,
            'winners': winners,
            'losers': total - winners,
            'win_rate': win_rate,
            'total_pnl': total_pnl,
            'avg_pnl': avg_pnl,
        }

    return results


# ============================================================================
# ANALYSIS FUNCTIONS
# ============================================================================

def calculate_trade_metrics(df: pd.DataFrame) -> Dict[str, Any]:
    """Calculate core trading metrics from MT5 history."""
    if df.empty:
        return {
            'total_trades': 0,
            'winners': 0,
            'losers': 0,
            'win_rate': 0.0,
            'profit_factor': 0.0,
            'total_pnl': 0.0,
            'gross_profit': 0.0,
            'gross_loss': 0.0,
            'avg_win': 0.0,
            'avg_loss': 0.0,
            'largest_win': 0.0,
            'largest_loss': 0.0,
            'avg_rr_actual': 0.0,
            'expectancy': 0.0,
            'max_drawdown': 0.0,
        }

    total_trades = len(df)
    winners = len(df[df['profit'] > 0])
    losers = len(df[df['profit'] < 0])
    breakeven = total_trades - winners - losers

    win_rate = (winners / total_trades * 100) if total_trades > 0 else 0

    gross_profit = df[df['profit'] > 0]['profit'].sum()
    gross_loss = abs(df[df['profit'] < 0]['profit'].sum())

    profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else float('inf')

    total_pnl = df['profit'].sum() + df['commission'].sum() + df['swap'].sum()

    avg_win = df[df['profit'] > 0]['profit'].mean() if winners > 0 else 0
    avg_loss = abs(df[df['profit'] < 0]['profit'].mean()) if losers > 0 else 0

    largest_win = df['profit'].max()
    largest_loss = df['profit'].min()

    # R-Multiple (avg win / avg loss)
    avg_rr_actual = (avg_win / avg_loss) if avg_loss > 0 else 0

    # Expectancy = (Win% x Avg Win) - (Loss% x Avg Loss)
    loss_rate = (losers / total_trades * 100) if total_trades > 0 else 0
    expectancy = (win_rate/100 * avg_win) - (loss_rate/100 * avg_loss)

    # Calculate drawdown
    cumulative = df['profit'].cumsum()
    peak = cumulative.cummax()
    drawdown = peak - cumulative
    max_drawdown = drawdown.max()

    return {
        'total_trades': total_trades,
        'winners': winners,
        'losers': losers,
        'breakeven': breakeven,
        'win_rate': win_rate,
        'profit_factor': profit_factor,
        'total_pnl': total_pnl,
        'gross_profit': gross_profit,
        'gross_loss': gross_loss,
        'avg_win': avg_win,
        'avg_loss': avg_loss,
        'largest_win': largest_win,
        'largest_loss': largest_loss,
        'avg_rr_actual': avg_rr_actual,
        'expectancy': expectancy,
        'max_drawdown': max_drawdown,
    }


def analyze_by_symbol(df: pd.DataFrame) -> Dict[str, Dict]:
    """Analyze performance by trading symbol."""
    results = {}

    if df.empty:
        return results

    for symbol in df['symbol'].unique():
        sym_df = df[df['symbol'] == symbol]
        metrics = calculate_trade_metrics(sym_df)
        results[symbol] = metrics

    return results


def analyze_by_hour(df: pd.DataFrame) -> Dict[int, Dict]:
    """Analyze performance by hour of day."""
    results = {}

    if df.empty or 'close_time' not in df.columns:
        return results

    df['hour'] = df['close_time'].apply(lambda x: x.hour if x else None)

    for hour in sorted(df['hour'].dropna().unique()):
        hour_df = df[df['hour'] == hour]
        metrics = calculate_trade_metrics(hour_df)
        results[int(hour)] = metrics

    return results


def analyze_by_day(df: pd.DataFrame) -> Dict[str, Dict]:
    """Analyze performance by day of week."""
    days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    results = {}

    if df.empty or 'close_time' not in df.columns:
        return results

    df['weekday'] = df['close_time'].apply(lambda x: x.weekday() if x else None)

    for day_num in range(7):
        day_df = df[df['weekday'] == day_num]
        if not day_df.empty:
            metrics = calculate_trade_metrics(day_df)
            results[days[day_num]] = metrics

    return results


def analyze_by_side(df: pd.DataFrame) -> Dict[str, Dict]:
    """Analyze performance by trade direction."""
    results = {}

    if df.empty or 'type' not in df.columns:
        return results

    for side in ['buy', 'sell']:
        side_df = df[df['type'].str.lower() == side]
        if not side_df.empty:
            metrics = calculate_trade_metrics(side_df)
            results[side.upper()] = metrics

    return results


def analyze_journal_signals(journal_trades: List[Dict]) -> Dict[str, Any]:
    """Analyze signal quality from journal entries."""
    if not journal_trades:
        return {
            'total_signals': 0,
            'avg_score': 0,
            'avg_confluence': 0,
            'score_distribution': {},
            'confluence_distribution': {},
        }

    scores = [t.get('score', 0) for t in journal_trades]
    confluences = [t.get('confluence', 0) for t in journal_trades]

    score_dist = defaultdict(int)
    for s in scores:
        bucket = int(s // 2) * 2  # Group by 2
        score_dist[f"{bucket}-{bucket+2}"] += 1

    conf_dist = defaultdict(int)
    for c in confluences:
        conf_dist[str(int(c))] += 1

    return {
        'total_signals': len(journal_trades),
        'avg_score': sum(scores) / len(scores) if scores else 0,
        'avg_confluence': sum(confluences) / len(confluences) if confluences else 0,
        'min_score': min(scores) if scores else 0,
        'max_score': max(scores) if scores else 0,
        'score_distribution': dict(score_dist),
        'confluence_distribution': dict(conf_dist),
    }


def analyze_by_symbol_journal(journal_trades: List[Dict]) -> Dict[str, Dict]:
    """Analyze signals by symbol from journal."""
    by_symbol = defaultdict(list)

    for trade in journal_trades:
        symbol = trade.get('symbol', 'UNKNOWN')
        by_symbol[symbol].append(trade)

    results = {}
    for symbol, trades in by_symbol.items():
        results[symbol] = {
            'count': len(trades),
            'avg_score': sum(t.get('score', 0) for t in trades) / len(trades) if trades else 0,
            'avg_confluence': sum(t.get('confluence', 0) for t in trades) / len(trades) if trades else 0,
            'long_count': len([t for t in trades if t.get('side', '').upper() == 'LONG']),
            'short_count': len([t for t in trades if t.get('side', '').upper() == 'SHORT']),
        }

    return results


def find_top_trades(df: pd.DataFrame, n: int = 5) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Find top winning and losing trades."""
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    top_wins = df.nlargest(n, 'profit')[['symbol', 'type', 'volume', 'profit', 'close_time']]
    top_losses = df.nsmallest(n, 'profit')[['symbol', 'type', 'volume', 'profit', 'close_time']]

    return top_wins, top_losses


def analyze_rejected_signals(audit_logs: List[Dict]) -> Dict[str, Any]:
    """Analyze rejected signals from audit logs."""
    rejected = [l for l in audit_logs if l.get('type') == 'SIGNAL_REJECTED']

    if not rejected:
        return {
            'total_rejected': 0,
            'reasons': {},
        }

    reasons = defaultdict(int)
    for r in rejected:
        reason = r.get('reason', 'unknown')
        reasons[reason] += 1

    return {
        'total_rejected': len(rejected),
        'reasons': dict(sorted(reasons.items(), key=lambda x: x[1], reverse=True)),
    }


# ============================================================================
# REPORT GENERATION
# ============================================================================

def calculate_health_score(metrics: Dict) -> float:
    """Calculate overall bot health score (0-10)."""
    score = 5.0  # Start neutral

    # Win rate contribution (target: 45%+)
    win_rate = metrics.get('win_rate', 0)
    if win_rate >= 55:
        score += 1.5
    elif win_rate >= 50:
        score += 1.0
    elif win_rate >= 45:
        score += 0.5
    elif win_rate < 40:
        score -= 1.0

    # Profit factor contribution (target: 1.3+)
    pf = metrics.get('profit_factor', 0)
    if pf >= 2.0:
        score += 1.5
    elif pf >= 1.5:
        score += 1.0
    elif pf >= 1.3:
        score += 0.5
    elif pf < 1.0:
        score -= 1.5

    # Expectancy contribution
    exp = metrics.get('expectancy', 0)
    if exp > 0:
        score += 1.0
    else:
        score -= 1.0

    # PnL contribution
    pnl = metrics.get('total_pnl', 0)
    if pnl > 0:
        score += 1.0
    else:
        score -= 0.5

    return min(10, max(0, score))


def generate_markdown_report(
    metrics: Dict,
    by_symbol: Dict,
    by_hour: Dict,
    by_day: Dict,
    by_side: Dict,
    signal_analysis: Dict,
    journal_by_symbol: Dict,
    top_wins: pd.DataFrame,
    top_losses: pd.DataFrame,
    rejected_analysis: Dict,
    start_date: datetime,
    end_date: datetime,
    by_exit_type: Optional[Dict] = None,
) -> str:
    """Generate the full Markdown report."""

    health_score = calculate_health_score(metrics)
    profitable = metrics.get('total_pnl', 0) > 0

    report = []

    # Header
    report.append("# RAPPORT HEBDOMADAIRE - EmpireAgentIA_3")
    report.append(f"\n**Periode:** {start_date.strftime('%Y-%m-%d')} au {end_date.strftime('%Y-%m-%d')}")
    report.append(f"**Genere le:** {datetime.now(LOCAL_TZ).strftime('%Y-%m-%d %H:%M:%S')}")
    report.append("")

    # 1. Executive Summary
    report.append("---")
    report.append("## 1. RESUME EXECUTIF")
    report.append("")
    report.append(f"- **Trades:** {metrics.get('total_trades', 0)} ({metrics.get('winners', 0)}W / {metrics.get('losers', 0)}L)")
    report.append(f"- **Win Rate:** {metrics.get('win_rate', 0):.1f}%")
    report.append(f"- **Profit Factor:** {metrics.get('profit_factor', 0):.2f}")
    report.append(f"- **P&L Total:** ${metrics.get('total_pnl', 0):,.2f}")
    report.append(f"- **Drawdown Max:** ${metrics.get('max_drawdown', 0):,.2f}")
    report.append("")
    verdict = "RENTABLE" if profitable else "DEFICITAIRE"
    report.append(f"**Verdict:** Semaine **{verdict}**")
    report.append(f"**Score de sante du bot:** {health_score:.1f}/10")
    report.append("")

    # 2. Detailed Statistics
    report.append("---")
    report.append("## 2. STATISTIQUES DETAILLEES")
    report.append("")
    report.append("### Performance globale")
    report.append("")
    report.append("| Metrique | Valeur |")
    report.append("|----------|--------|")
    report.append(f"| Trades totaux | {metrics.get('total_trades', 0)} |")
    report.append(f"| Gagnants | {metrics.get('winners', 0)} |")
    report.append(f"| Perdants | {metrics.get('losers', 0)} |")
    report.append(f"| Win Rate | {metrics.get('win_rate', 0):.1f}% |")
    report.append(f"| Profit Factor | {metrics.get('profit_factor', 0):.2f} |")
    report.append(f"| Gain brut | ${metrics.get('gross_profit', 0):,.2f} |")
    report.append(f"| Perte brute | -${metrics.get('gross_loss', 0):,.2f} |")
    report.append(f"| P&L Net | ${metrics.get('total_pnl', 0):,.2f} |")
    report.append(f"| Gain moyen | ${metrics.get('avg_win', 0):,.2f} |")
    report.append(f"| Perte moyenne | -${metrics.get('avg_loss', 0):,.2f} |")
    report.append(f"| Meilleur trade | ${metrics.get('largest_win', 0):,.2f} |")
    report.append(f"| Pire trade | ${metrics.get('largest_loss', 0):,.2f} |")
    report.append(f"| R-Multiple moyen | {metrics.get('avg_rr_actual', 0):.2f} |")
    report.append(f"| Esperance | ${metrics.get('expectancy', 0):,.2f} |")
    report.append(f"| Drawdown max | ${metrics.get('max_drawdown', 0):,.2f} |")
    report.append("")

    # By Symbol
    if by_symbol:
        report.append("### Performance par symbole")
        report.append("")
        report.append("| Symbole | Trades | Win% | PF | P&L |")
        report.append("|---------|--------|------|-----|-----|")
        sorted_symbols = sorted(by_symbol.items(), key=lambda x: x[1].get('total_pnl', 0), reverse=True)
        for symbol, data in sorted_symbols:
            report.append(f"| {symbol} | {data.get('total_trades', 0)} | {data.get('win_rate', 0):.1f}% | {data.get('profit_factor', 0):.2f} | ${data.get('total_pnl', 0):,.2f} |")
        report.append("")

    # By Day
    if by_day:
        report.append("### Performance par jour")
        report.append("")
        report.append("| Jour | Trades | Win% | P&L |")
        report.append("|------|--------|------|-----|")
        for day, data in by_day.items():
            report.append(f"| {day} | {data.get('total_trades', 0)} | {data.get('win_rate', 0):.1f}% | ${data.get('total_pnl', 0):,.2f} |")
        report.append("")

    # By Hour
    if by_hour:
        report.append("### Performance par heure (Top 5)")
        report.append("")
        report.append("| Heure | Trades | Win% | P&L |")
        report.append("|-------|--------|------|-----|")
        sorted_hours = sorted(by_hour.items(), key=lambda x: x[1].get('total_pnl', 0), reverse=True)[:5]
        for hour, data in sorted_hours:
            report.append(f"| {hour:02d}:00 | {data.get('total_trades', 0)} | {data.get('win_rate', 0):.1f}% | ${data.get('total_pnl', 0):,.2f} |")
        report.append("")

    # By Side
    if by_side:
        report.append("### Performance par direction")
        report.append("")
        report.append("| Direction | Trades | Win% | P&L |")
        report.append("|-----------|--------|------|-----|")
        for side, data in by_side.items():
            report.append(f"| {side} | {data.get('total_trades', 0)} | {data.get('win_rate', 0):.1f}% | ${data.get('total_pnl', 0):,.2f} |")
        report.append("")

    # By Exit Type
    if by_exit_type:
        report.append("### Performance par type de sortie")
        report.append("")
        report.append("| Type | Trades | Win% | P&L | Avg P&L |")
        report.append("|------|--------|------|-----|---------|")

        # Labels lisibles pour chaque type
        exit_labels = {
            "tp": "Take Profit",
            "sl": "Stop Loss",
            "be": "Break Even",
            "trailing": "Trailing Stop",
            "partial": "Partial Close",
            "manual": "MANUEL",
            "unknown": "Inconnu"
        }

        for exit_type, data in sorted(by_exit_type.items(), key=lambda x: x[1].get('trades', 0), reverse=True):
            label = exit_labels.get(exit_type, exit_type)
            # Mettre en evidence les trades manuels
            if exit_type == "manual":
                label = f"**{label}**"
            report.append(
                f"| {label} | {data.get('trades', 0)} | {data.get('win_rate', 0):.1f}% | "
                f"${data.get('total_pnl', 0):,.2f} | ${data.get('avg_pnl', 0):,.2f} |"
            )
        report.append("")

        # Avertissement si beaucoup de trades manuels
        manual_data = by_exit_type.get("manual", {})
        if manual_data.get("trades", 0) > 0:
            total_trades = sum(d.get("trades", 0) for d in by_exit_type.values())
            manual_pct = (manual_data.get("trades", 0) / total_trades * 100) if total_trades > 0 else 0
            if manual_pct > 20:
                report.append(f"> **ATTENTION:** {manual_pct:.0f}% des trades ont ete fermes manuellement. ")
                report.append("> Cela peut fausser l'evaluation de la strategie automatique.")
                report.append("")

    # Signal Analysis
    if signal_analysis.get('total_signals', 0) > 0:
        report.append("### Qualite des signaux")
        report.append("")
        report.append(f"- **Signaux emis:** {signal_analysis.get('total_signals', 0)}")
        report.append(f"- **Score moyen:** {signal_analysis.get('avg_score', 0):.1f}")
        report.append(f"- **Confluence moyenne:** {signal_analysis.get('avg_confluence', 0):.1f}")
        report.append(f"- **Score min/max:** {signal_analysis.get('min_score', 0):.1f} / {signal_analysis.get('max_score', 0):.1f}")
        report.append("")

    # 3. Top Trades
    report.append("---")
    report.append("## 3. TOP TRADES")
    report.append("")

    if not top_wins.empty:
        report.append("### Top 5 meilleurs trades")
        report.append("")
        report.append("| # | Symbole | Type | Volume | Profit | Date |")
        report.append("|---|---------|------|--------|--------|------|")
        for idx, (_, row) in enumerate(top_wins.iterrows(), 1):
            date_str = row['close_time'].strftime('%Y-%m-%d %H:%M') if row['close_time'] else ''
            report.append(f"| {idx} | {row['symbol']} | {row['type']} | {row['volume']:.2f} | ${row['profit']:,.2f} | {date_str} |")
        report.append("")

    if not top_losses.empty:
        report.append("### Top 5 pires trades")
        report.append("")
        report.append("| # | Symbole | Type | Volume | Perte | Date |")
        report.append("|---|---------|------|--------|-------|------|")
        for idx, (_, row) in enumerate(top_losses.iterrows(), 1):
            date_str = row['close_time'].strftime('%Y-%m-%d %H:%M') if row['close_time'] else ''
            report.append(f"| {idx} | {row['symbol']} | {row['type']} | {row['volume']:.2f} | ${row['profit']:,.2f} | {date_str} |")
        report.append("")

    # 4. Strengths
    report.append("---")
    report.append("## 4. POINTS FORTS")
    report.append("")

    strengths = []

    if metrics.get('win_rate', 0) >= 50:
        strengths.append(f"Win rate superieur a 50% ({metrics.get('win_rate', 0):.1f}%)")

    if metrics.get('profit_factor', 0) >= 1.5:
        strengths.append(f"Excellent Profit Factor ({metrics.get('profit_factor', 0):.2f})")

    if metrics.get('expectancy', 0) > 0:
        strengths.append(f"Esperance positive (${metrics.get('expectancy', 0):.2f} par trade)")

    # Best performing symbols
    if by_symbol:
        best_symbols = [s for s, d in by_symbol.items() if d.get('win_rate', 0) >= 55 and d.get('total_trades', 0) >= 3]
        if best_symbols:
            strengths.append(f"Symboles performants: {', '.join(best_symbols)}")

    if by_side:
        for side, data in by_side.items():
            if data.get('win_rate', 0) >= 55:
                strengths.append(f"Direction {side} performante ({data.get('win_rate', 0):.1f}% win rate)")

    if strengths:
        for s in strengths:
            report.append(f"- {s}")
    else:
        report.append("- Aucun point fort identifie cette semaine")
    report.append("")

    # 5. Weaknesses
    report.append("---")
    report.append("## 5. POINTS FAIBLES")
    report.append("")

    weaknesses = []

    if metrics.get('win_rate', 0) < 45:
        weaknesses.append(f"Win rate insuffisant ({metrics.get('win_rate', 0):.1f}% < 45%)")

    if metrics.get('profit_factor', 0) < 1.3 and metrics.get('profit_factor', 0) > 0:
        weaknesses.append(f"Profit Factor faible ({metrics.get('profit_factor', 0):.2f} < 1.3)")

    if metrics.get('expectancy', 0) <= 0:
        weaknesses.append(f"Esperance negative (${metrics.get('expectancy', 0):.2f})")

    # Worst performing symbols
    if by_symbol:
        worst_symbols = [(s, d) for s, d in by_symbol.items() if d.get('total_pnl', 0) < 0 and d.get('total_trades', 0) >= 3]
        if worst_symbols:
            worst_names = [s for s, _ in sorted(worst_symbols, key=lambda x: x[1].get('total_pnl', 0))[:3]]
            weaknesses.append(f"Symboles deficitaires: {', '.join(worst_names)}")

    if by_side:
        for side, data in by_side.items():
            if data.get('win_rate', 0) < 40 and data.get('total_trades', 0) >= 5:
                weaknesses.append(f"Direction {side} sous-performante ({data.get('win_rate', 0):.1f}% win rate)")

    if weaknesses:
        for w in weaknesses:
            report.append(f"- {w}")
    else:
        report.append("- Aucune faiblesse majeure identifiee")
    report.append("")

    # 6. Recommendations
    report.append("---")
    report.append("## 6. RECOMMANDATIONS")
    report.append("")

    recommendations = []

    # Win rate recommendations
    if metrics.get('win_rate', 0) < 45:
        recommendations.append({
            'problem': f"Win rate trop bas ({metrics.get('win_rate', 0):.1f}%)",
            'cause': "Signaux de qualite insuffisante ou entrees mal timees",
            'actions': [
                "Augmenter min_score de 5.0 a 7.0",
                "Augmenter min_confluence de 3 a 4",
                "Revoir les conditions d'entree des agents"
            ],
            'file': "config/config.yaml",
        })

    # Profit factor recommendations
    if 0 < metrics.get('profit_factor', 0) < 1.3:
        recommendations.append({
            'problem': f"Profit Factor faible ({metrics.get('profit_factor', 0):.2f})",
            'cause': "Ratio risque/reward insuffisant ou gains coupes trop tot",
            'actions': [
                "Augmenter le target R/R de 1.5 a 2.0",
                "Implementer un trailing stop plus agressif",
                "Reduire le stop loss pour un meilleur R/R"
            ],
            'file': "config/config.yaml",
        })

    # Symbol-specific recommendations
    if by_symbol:
        for symbol, data in by_symbol.items():
            if data.get('total_trades', 0) >= 5 and data.get('win_rate', 0) < 35:
                recommendations.append({
                    'problem': f"{symbol}: Win rate critique ({data.get('win_rate', 0):.1f}%)",
                    'cause': f"Ce symbole ne correspond pas aux strategies actuelles",
                    'actions': [
                        f"Desactiver temporairement {symbol}",
                        "Analyser les conditions de marche specifiques",
                        "Revoir les parametres specifiques au symbole"
                    ],
                    'file': "config/symbols.yaml",
                })

    # Hour-based recommendations
    if by_hour:
        worst_hours = [(h, d) for h, d in by_hour.items() if d.get('total_pnl', 0) < 0 and d.get('total_trades', 0) >= 3]
        if worst_hours:
            worst_hours_list = sorted(worst_hours, key=lambda x: x[1].get('total_pnl', 0))[:3]
            hours_str = ', '.join([f"{h:02d}:00" for h, _ in worst_hours_list])
            recommendations.append({
                'problem': f"Heures deficitaires: {hours_str}",
                'cause': "Sessions de trading non optimales",
                'actions': [
                    "Revoir les heures de trading autorisees",
                    "Reduire l'exposition pendant ces heures",
                    "Ajouter des filtres de session"
                ],
                'file': "config/config.yaml",
            })

    if recommendations:
        for i, rec in enumerate(recommendations, 1):
            report.append(f"### Recommandation {i}")
            report.append("")
            report.append(f"**PROBLEME:** {rec['problem']}")
            report.append(f"**CAUSE PROBABLE:** {rec['cause']}")
            report.append("**ACTIONS:**")
            for action in rec['actions']:
                report.append(f"- {action}")
            report.append(f"**FICHIER:** `{rec['file']}`")
            report.append("")
    else:
        report.append("Aucune recommandation majeure - Performance satisfaisante.")
        report.append("")

    # 7. Parameter Adjustments
    report.append("---")
    report.append("## 7. PARAMETRES A AJUSTER")
    report.append("")

    current_min_score = 5.0  # Default assumed
    current_min_confluence = 3  # Default assumed

    suggested_min_score = current_min_score
    suggested_min_confluence = current_min_confluence

    if metrics.get('win_rate', 0) < 45:
        suggested_min_score = max(7.0, current_min_score + 2)
        suggested_min_confluence = max(4, current_min_confluence + 1)
    elif metrics.get('win_rate', 0) < 50:
        suggested_min_score = max(6.0, current_min_score + 1)

    report.append("| Parametre | Actuel | Recommande | Raison |")
    report.append("|-----------|--------|------------|--------|")
    report.append(f"| min_score | {current_min_score} | {suggested_min_score} | {'Ameliorer filtrage' if suggested_min_score > current_min_score else 'OK'} |")
    report.append(f"| min_confluence | {current_min_confluence} | {suggested_min_confluence} | {'Plus de confirmation' if suggested_min_confluence > current_min_confluence else 'OK'} |")

    # Symbols to disable
    if by_symbol:
        disable_symbols = [s for s, d in by_symbol.items() if d.get('total_trades', 0) >= 5 and d.get('win_rate', 0) < 35]
        if disable_symbols:
            report.append(f"| symbols_to_disable | - | {', '.join(disable_symbols)} | Performance critique |")

    report.append("")

    # 8. Priority Action Plan
    report.append("---")
    report.append("## 8. PLAN D'ACTION PRIORITAIRE")
    report.append("")

    actions = []

    if metrics.get('win_rate', 0) < 45:
        actions.append("1. **URGENT:** Augmenter les seuils de filtrage (min_score, min_confluence)")

    if metrics.get('profit_factor', 0) < 1.0:
        actions.append("2. **CRITIQUE:** Revoir la strategie de sortie (R/R target, trailing stop)")

    if by_symbol:
        worst = [(s, d) for s, d in by_symbol.items() if d.get('total_pnl', 0) < -100]
        if worst:
            symbols = ', '.join([s for s, _ in sorted(worst, key=lambda x: x[1].get('total_pnl', 0))[:2]])
            actions.append(f"3. **IMPORTANT:** Desactiver temporairement les symboles deficitaires: {symbols}")

    if not actions:
        actions.append("Performance satisfaisante - Continuer le monitoring")

    for action in actions:
        report.append(action)
    report.append("")

    # Rejected signals
    if rejected_analysis.get('total_rejected', 0) > 0:
        report.append("---")
        report.append("## 9. SIGNAUX REJETES")
        report.append("")
        report.append(f"**Total rejetes:** {rejected_analysis.get('total_rejected', 0)}")
        report.append("")
        report.append("**Raisons principales:**")
        for reason, count in list(rejected_analysis.get('reasons', {}).items())[:5]:
            report.append(f"- {reason}: {count}")
        report.append("")

    return "\n".join(report)


def generate_telegram_summary(metrics: Dict, start_date: datetime, end_date: datetime) -> str:
    """Generate a short summary for Telegram (max 500 chars)."""
    health_score = calculate_health_score(metrics)
    profitable = metrics.get('total_pnl', 0) > 0
    emoji = "+" if profitable else "-"
    status = "RENTABLE" if profitable else "DEFICIT"

    summary = f"""RAPPORT SEMAINE {start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m')}

Trades: {metrics.get('total_trades', 0)} ({metrics.get('winners', 0)}W/{metrics.get('losers', 0)}L)
Win Rate: {metrics.get('win_rate', 0):.1f}%
PF: {metrics.get('profit_factor', 0):.2f}
P&L: {emoji}${abs(metrics.get('total_pnl', 0)):,.0f}

Verdict: {status}
Score: {health_score:.1f}/10"""

    return summary[:500]


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Analyze weekly trading performance")
    parser.add_argument("--days", type=int, default=7, help="Number of days to analyze")
    parser.add_argument("--output", type=str, default="reports/weekly_report.md", help="Output file path")
    args = parser.parse_args()

    # Date range
    end_date = datetime.now(LOCAL_TZ)
    start_date = end_date - timedelta(days=args.days)

    print(f"[REPORT] Analyzing period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    # Paths
    project_root = ROOT
    data_dir = project_root / "data"
    logs_dir = project_root / "logs"
    reports_dir = project_root / "reports"
    excel_path = reports_dir / "ReportHistory-10960352.xlsx"

    # Load data
    print("[REPORT] Loading MT5 history from Excel...")
    mt5_df = load_mt5_history_from_excel(excel_path, start_date, end_date)
    print(f"[REPORT] Loaded {len(mt5_df)} trades from MT5")

    print("[REPORT] Loading journal trades...")
    journal_trades = load_journal_trades(data_dir, start_date, end_date)
    print(f"[REPORT] Loaded {len(journal_trades)} journal entries")

    print("[REPORT] Loading audit logs...")
    audit_logs = load_audit_logs(logs_dir, start_date, end_date)
    print(f"[REPORT] Loaded {len(audit_logs)} audit entries")

    print("[REPORT] Loading trade outcomes with exit types...")
    outcomes_df = load_trade_outcomes(data_dir, start_date, end_date)
    print(f"[REPORT] Loaded {len(outcomes_df)} trade outcomes")

    # Analysis
    print("[REPORT] Calculating metrics...")
    metrics = calculate_trade_metrics(mt5_df)
    by_symbol = analyze_by_symbol(mt5_df)
    by_hour = analyze_by_hour(mt5_df)
    by_day = analyze_by_day(mt5_df)
    by_side = analyze_by_side(mt5_df)
    signal_analysis = analyze_journal_signals(journal_trades)
    journal_by_symbol = analyze_by_symbol_journal(journal_trades)
    top_wins, top_losses = find_top_trades(mt5_df)
    rejected_analysis = analyze_rejected_signals(audit_logs)
    by_exit_type = analyze_by_exit_type(outcomes_df)

    # Generate report
    print("[REPORT] Generating Markdown report...")
    report = generate_markdown_report(
        metrics, by_symbol, by_hour, by_day, by_side,
        signal_analysis, journal_by_symbol,
        top_wins, top_losses, rejected_analysis,
        start_date, end_date,
        by_exit_type=by_exit_type
    )

    # Save report
    output_path = project_root / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(report, encoding='utf-8')
    print(f"[REPORT] Report saved to: {output_path}")

    # Generate Telegram summary
    telegram_summary = generate_telegram_summary(metrics, start_date, end_date)
    telegram_path = output_path.with_suffix('.telegram.txt')
    telegram_path.write_text(telegram_summary, encoding='utf-8')
    print(f"[REPORT] Telegram summary saved to: {telegram_path}")

    # Print summary
    print("\n" + "="*60)
    print("RESUME")
    print("="*60)
    print(f"Trades: {metrics.get('total_trades', 0)}")
    print(f"Win Rate: {metrics.get('win_rate', 0):.1f}%")
    print(f"Profit Factor: {metrics.get('profit_factor', 0):.2f}")
    print(f"P&L: ${metrics.get('total_pnl', 0):,.2f}")
    print(f"Health Score: {calculate_health_score(metrics):.1f}/10")
    print("="*60)

    return 0


if __name__ == "__main__":
    sys.exit(main())
